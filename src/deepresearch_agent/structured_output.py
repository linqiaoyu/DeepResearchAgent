from __future__ import annotations

import csv
import importlib.util
import io
import json
import re
from datetime import date, datetime, timezone
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

from deepresearch_agent.schemas import (
    ComparisonTable,
    Evidence,
    EventTimeline,
    MetricRow,
    ResearchState,
    RiskItem,
    RiskMatrix,
    StructuredResearchOutput,
    TimelineEvent,
)
from deepresearch_agent.domains.registry import load_domain_pack

_FIXED_WORKBOOK_TIME = datetime(2026, 7, 9, tzinfo=timezone.utc)
_FIXED_ZIP_TIME = (2026, 7, 9, 0, 0, 0)
_MODIFIED_PROPERTY_RE = re.compile(
    rb"(<dcterms:modified\b[^>]*>)[^<]*(</dcterms:modified>)"
)
_ENGLISH_METRICS = (
    (
        re.compile(r"Advisor productivity improved (?P<value>\d+(?:\.\d+)?)%"),
        "wealth management pilot",
        "Advisor productivity",
        "advisor productivity",
        "pilot cohort",
    ),
    (
        re.compile(
            r"Assets under management growth remained (?P<value>\d+(?:\.\d+)?)%"
        ),
        "digital wealth cohort",
        "Assets under management growth",
        "assets under management growth",
        "surveyed cohort",
    ),
)


def build_structured_output(state: ResearchState) -> StructuredResearchOutput:
    aliases = _metric_aliases()
    data_as_of = max(
        (item.source_pub_date for item in state.evidence_store if item.source_pub_date),
        default=None,
    )
    metric_rows = [
        row
        for item in state.evidence_store
        for row in _metric_rows_for_evidence(item, aliases)
    ]
    metric_rows.sort(
        key=lambda row: (
            row.entity,
            row.normalized_metric,
            row.period,
            row.scope,
            row.value,
            row.evidence_ids,
        )
    )
    scope_notes = _scope_notes(metric_rows)

    events = [
        TimelineEvent(
            occurred_at=item.source_pub_date,
            event=item.claim,
            source=item.source_url,
            thesis_impact=_thesis_impact(item.claim_type, item.confidence),
            evidence_ids=[item.id],
            verification_status=(
                "unverified" if item.claim_type == "projection" else "verified"
            ),
        )
        for item in state.evidence_store
    ]
    events.sort(key=lambda row: (row.occurred_at is None, row.occurred_at or date.max, row.event, row.source, row.evidence_ids))

    evidence_by_claim = {item.claim: item.id for item in state.evidence_store}
    risks: list[RiskItem] = []
    if state.critic_report:
        for issue in state.critic_report.issues:
            evidence_ids = sorted(
                {
                    evidence_by_claim[claim]
                    for claim in issue.affected_claims
                    if claim in evidence_by_claim
                }
            )
            risks.append(
                RiskItem(
                    risk=issue.message,
                    likelihood="unknown",
                    impact=issue.severity,
                    evidence_ids=evidence_ids,
                    verification_status=(
                        "verified" if evidence_ids else "unverified"
                    ),
                    unverified_prediction=issue.issue_type == "unverified_projection",
                )
            )
    risks.sort(key=lambda row: (row.impact, row.risk, row.evidence_ids))

    return StructuredResearchOutput(
        comparison_table=ComparisonTable(
            question=state.topic,
            rows=metric_rows,
            scope_consistent=not scope_notes,
            scope_notes=scope_notes,
            data_as_of=data_as_of,
        ),
        event_timeline=EventTimeline(
            question=state.topic,
            events=events,
            data_as_of=data_as_of,
        ),
        risk_matrix=RiskMatrix(
            question=state.topic,
            risks=risks,
            data_as_of=data_as_of,
        ),
    )


def metric_fact_keys(
    evidence: list[Evidence],
) -> dict[str, set[tuple[str, str, str, str]]]:
    """Return reader-deduplication keys without changing structured values."""

    aliases = _metric_aliases()
    return {
        item.id: {
            (
                row.entity,
                row.normalized_metric,
                (
                    row.period[:4]
                    if re.fullmatch(r"\d{4}1231", row.period)
                    else row.period
                ),
                row.scope,
            )
            for row in _metric_rows_for_evidence(item, aliases)
        }
        for item in evidence
    }


def render_structured_json(output: StructuredResearchOutput) -> str:
    return json.dumps(
        output.model_dump(mode="json"),
        ensure_ascii=False,
        indent=2,
        sort_keys=True,
    ) + "\n"


def render_structured_markdown(output: StructuredResearchOutput) -> str:
    table = output.comparison_table
    lines = [
        "## 结构化指标对比",
        "",
        "| Entity | Metric | Normalized metric | Period | Scope | Value | Unit | Evidence | Confidence | Status |",
        "| --- | --- | --- | --- | --- | ---: | --- | --- | ---: | --- |",
    ]
    for row in table.rows:
        lines.append(
            "| "
            + " | ".join(
                [
                    _md(row.entity),
                    _md(row.metric),
                    _md(row.normalized_metric),
                    _md(row.period),
                    _md(row.scope),
                    _number(row.value),
                    _md(row.unit),
                    _md(", ".join(row.evidence_ids)),
                    _number(row.confidence),
                    row.verification_status,
                ]
            )
            + " |"
        )
    if not table.rows:
        lines.append("| — | — | — | — | — | — | — | — | — | unverified |")
    lines.extend(
        [
            "",
            f"口径一致性：{'通过' if table.scope_consistent else '存在冲突'}",
        ]
    )
    lines.extend(f"- 表注：{note}" for note in table.scope_notes)
    lines.extend(
        [
            "",
            "## 事件时间线",
            "",
            "| Date | Event | Source | Impact | Evidence | Status |",
            "| --- | --- | --- | --- | --- | --- |",
        ]
    )
    for row in output.event_timeline.events:
        lines.append(
            "| "
            + " | ".join(
                [
                    row.occurred_at.isoformat() if row.occurred_at else "unknown",
                    _md(row.event),
                    _md(row.source),
                    row.thesis_impact,
                    _md(", ".join(row.evidence_ids)),
                    row.verification_status,
                ]
            )
            + " |"
        )
    lines.extend(
        [
            "",
            "## 风险矩阵",
            "",
            "| Risk | Likelihood | Impact | Evidence | Unverified prediction | Status |",
            "| --- | --- | --- | --- | --- | --- |",
        ]
    )
    for row in output.risk_matrix.risks:
        lines.append(
            "| "
            + " | ".join(
                [
                    _md(row.risk),
                    row.likelihood,
                    row.impact,
                    _md(", ".join(row.evidence_ids)),
                    "yes" if row.unverified_prediction else "no",
                    row.verification_status,
                ]
            )
            + " |"
        )
    return "\n".join(lines) + "\n"


def write_structured_table(
    output: StructuredResearchOutput,
    path: Path,
) -> Path:
    if importlib.util.find_spec("openpyxl") is not None:
        actual_path = path.with_suffix(".xlsx")
        actual_path.parent.mkdir(parents=True, exist_ok=True)
        actual_path.write_bytes(_xlsx_bytes(output))
        return actual_path
    actual_path = path.with_suffix(".csv")
    actual_path.parent.mkdir(parents=True, exist_ok=True)
    with actual_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle, lineterminator="\n")
        writer.writerow(_metric_headers())
        for row in output.comparison_table.rows:
            writer.writerow(_metric_values(row))
    return actual_path


def _xlsx_bytes(output: StructuredResearchOutput) -> bytes:
    from openpyxl import Workbook

    workbook = Workbook()
    workbook.properties.created = _FIXED_WORKBOOK_TIME
    workbook.properties.modified = _FIXED_WORKBOOK_TIME
    metric_sheet = workbook.active
    metric_sheet.title = "metrics"
    metric_sheet.append(_metric_headers())
    for row in output.comparison_table.rows:
        metric_sheet.append(_metric_values(row))

    timeline_sheet = workbook.create_sheet("timeline")
    timeline_sheet.append(["occurred_at", "event", "source", "impact", "evidence_ids", "status"])
    for row in output.event_timeline.events:
        timeline_sheet.append(
            [
                row.occurred_at.isoformat() if row.occurred_at else "unknown",
                row.event,
                row.source,
                row.thesis_impact,
                ",".join(row.evidence_ids),
                row.verification_status,
            ]
        )

    risk_sheet = workbook.create_sheet("risks")
    risk_sheet.append(
        ["risk", "likelihood", "impact", "evidence_ids", "unverified_prediction", "status"]
    )
    for row in output.risk_matrix.risks:
        risk_sheet.append(
            [
                row.risk,
                row.likelihood,
                row.impact,
                ",".join(row.evidence_ids),
                row.unverified_prediction,
                row.verification_status,
            ]
        )

    raw = io.BytesIO()
    workbook.save(raw)
    normalized = io.BytesIO()
    with ZipFile(io.BytesIO(raw.getvalue()), "r") as source:
        with ZipFile(normalized, "w", compression=ZIP_DEFLATED, compresslevel=9) as target:
            for name in sorted(source.namelist()):
                info = ZipInfo(name, date_time=_FIXED_ZIP_TIME)
                info.compress_type = ZIP_DEFLATED
                info.external_attr = source.getinfo(name).external_attr
                content = source.read(name)
                if name == "docProps/core.xml":
                    content = _MODIFIED_PROPERTY_RE.sub(
                        rb"\g<1>2026-07-09T00:00:00Z\g<2>",
                        content,
                    )
                target.writestr(info, content)
    return normalized.getvalue()


def _metric_aliases() -> dict[str, str]:
    path = load_domain_pack("finance").metric_table_path()
    payload = json.loads(path.read_text(encoding="utf-8"))
    return {str(key): str(value) for key, value in payload["metric_aliases"].items()}


def _metric_rows_for_evidence(
    item: Evidence,
    aliases: dict[str, str],
) -> list[MetricRow]:
    fields = item.numeric_fields
    if fields is not None and fields.value is not None:
        metric = fields.metric_name or "未标注"
        return [
            MetricRow(
                entity=fields.entity or "未标注",
                metric=metric,
                normalized_metric=aliases.get(metric, metric),
                period=fields.period or "未标注",
                scope=fields.dimension or "未标注",
                value=float(fields.value),
                unit=fields.unit or "未标注",
                evidence_ids=[item.id],
                confidence=item.confidence,
                verification_status="verified",
            )
        ]
    if item.claim_type != "data":
        return []
    rows: list[MetricRow] = []
    for match in load_domain_pack("finance").metric_claim_pattern().finditer(item.claim):
        metric = match.group("metric")
        rows.append(
            MetricRow(
                entity=match.group("entity"),
                metric=metric,
                normalized_metric=aliases.get(metric, metric),
                period=match.group("period"),
                scope=match.group("scope") or "未标注",
                value=float(match.group("value")),
                unit=match.group("unit"),
                evidence_ids=[item.id],
                confidence=item.confidence,
                verification_status="verified",
            )
        )
    for pattern, entity, metric, normalized_metric, scope in _ENGLISH_METRICS:
        match = pattern.search(item.claim)
        if match:
            rows.append(
                MetricRow(
                    entity=entity,
                    metric=metric,
                    normalized_metric=normalized_metric,
                    period="未标注",
                    scope=scope,
                    value=float(match.group("value")),
                    unit="%",
                    evidence_ids=[item.id],
                    confidence=item.confidence,
                    verification_status="verified",
                )
            )
    return rows


def _scope_notes(rows: list[MetricRow]) -> list[str]:
    scopes: dict[str, set[str]] = {}
    for row in rows:
        scopes.setdefault(row.normalized_metric, set()).add(row.scope)
    return [
        f"{metric} 同时存在不同口径：{', '.join(sorted(values))}；禁止静默并列。"
        for metric, values in sorted(scopes.items())
        if len(values) > 1
    ]


def _thesis_impact(claim_type: str, confidence: float) -> str:
    if claim_type == "projection":
        return "uncertain"
    if claim_type == "opinion" or confidence < 0.6:
        return "neutral"
    return "positive"


def _metric_headers() -> list[str]:
    return [
        "entity",
        "metric",
        "normalized_metric",
        "period",
        "scope",
        "value",
        "unit",
        "evidence_ids",
        "confidence",
        "status",
    ]


def _metric_values(row: MetricRow) -> list[str | float]:
    return [
        row.entity,
        row.metric,
        row.normalized_metric,
        row.period,
        row.scope,
        row.value,
        row.unit,
        ",".join(row.evidence_ids),
        row.confidence,
        row.verification_status,
    ]


def _md(value: str) -> str:
    return value.replace("|", "\\|").replace("\n", " ")


def _number(value: float) -> str:
    return format(value, ".12g")
